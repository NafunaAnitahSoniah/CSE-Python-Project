from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, F, Q
from django.http import HttpResponse, HttpResponseBadRequest
from datetime import datetime
from django.utils import timezone
from django.core.exceptions import PermissionDenied
from .models import *
from .forms import *
from django.contrib.auth import authenticate, login, logout
from django.db import transaction
from django.contrib.auth.forms import AuthenticationForm
from django.db import transaction


# Create your views here.
# Landing page
def indexpage(request):
    # Get search query and items per page from request
    search_query = request.GET.get('q', '')
    items_per_page = int(request.GET.get('per_page', 10))

    # Get all chick requests, ordered by request date
    chick_requests_qs = ChickRequest.objects.select_related('farmer').order_by('-request_date')

    # Apply search filter
    if search_query:
        chick_requests_qs = chick_requests_qs.filter(
            Q(farmer__farmer_name__icontains=search_query) | 
            Q(chick_request_id__icontains=search_query)
        )

    # Paginate the results
    from django.core.paginator import Paginator
    paginator = Paginator(chick_requests_qs, items_per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Create selection flags for the dropdown
    is_10_selected = items_per_page == 10
    is_20_selected = items_per_page == 20
    is_100_selected = items_per_page == 100

    return render(request, 'index.html', {
        'page_obj': page_obj,
        'search_query': search_query,
        'items_per_page': items_per_page,
        'is_10_selected': is_10_selected,
        'is_20_selected': is_20_selected,
        'is_100_selected': is_100_selected,
    })

def signup(request):
    if request.method == 'POST':
        #creating access to our form(forms.py)
        form_data = UserCreation(request.POST)
        #to check if the data posted from the UserCreation form is valid 
        if form_data.is_valid():
            #the validated data is then saved in the database table called Farmer_users
            form_data.save() 
            #directs django to pay attention to the fields username and email; and clean them
            username = form_data.cleaned_data.get('username')
            email = form_data.cleaned_data.get('email')
            #so after successfully registering, you are directed to the login form/ page
            return redirect('login')
    #in case there is no data coming from the UserCreation form, direct to the signup.html page   
    else:
        form_data = UserCreation()
    return render(request, 'signup.html', {'form_data': form_data})

def Login(request):
    # If already authenticated send to their dashboard
    if request.user.is_authenticated:
        return _redirect_by_role(request.user)

    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user:
                login(request, user)
                return _redirect_by_role(user)
        # invalid: surface error message
        messages.error(request, 'Invalid username or password.')

    return render(request, 'login.html', {'form': form, 'title': 'Login'})


def _redirect_by_role(user):
    if getattr(user, 'role', None) == 'sales_agent':
        return redirect('salesagentdashboard')
    if getattr(user, 'role', None) == 'manager':
        return redirect('Managersdashboard')
    return redirect('/')


def role_required(role):
    def decorator(view_func):
        @login_required
        def _wrapped(request, *args, **kwargs):
            if getattr(request.user, 'role', None) != role:
                raise PermissionDenied
            return view_func(request, *args, **kwargs)
        return _wrapped
    return decorator


#For the Managers

@role_required('manager')
def Managersdashboard(request):
    total_users = UserProfile.objects.count()
    chick_stock = ChickStock.objects.aggregate(total=Sum('stock_quantity'))['total'] or 0
    feed_stock = FeedStock.objects.aggregate(total=Sum('feed_quantity'))['total'] or 0
    # Combine chick and feed request stats
    chick_stats = ChickRequest.objects.values('status').annotate(c=Count('id'))
    feed_stats = FeedAllocation.objects.values('status').annotate(c=Count('id'))
    chick_map = {r['status']: r['c'] for r in chick_stats}
    feed_map = {r['status']: r['c'] for r in feed_stats}
    pending_requests = chick_map.get('pending', 0) + feed_map.get('pending', 0)
    approved_requests = chick_map.get('approved', 0) + feed_map.get('approved', 0)
    rejected_requests = chick_map.get('rejected', 0) + feed_map.get('rejected', 0)
    # Total sales amount = feed (approved: bags*selling_price) + chicks (approved: qty*price)
    feed_total = FeedAllocation.objects.filter(status='approved').aggregate(
        total=Sum(F('bags_allocated') * F('feed_stock__selling_price'))
    )['total'] or 0
    # Latest price per type/breed: use the most recent updated stock entry
    # For aggregate simplicity, iterate in Python (small datasets typical for dashboard)
    chick_total = 0
    for req in ChickRequest.objects.filter(status='approved').select_related('farmer'):
        price = ChickStock.objects.filter(
            chick_type=req.chick_type,
            chick_breed=req.chick_breed
        ).order_by('-updated_at').values_list('chick_price', flat=True).first() or 1650
        chick_total += int(req.quantity or 0) * int(price)
    total_sales = feed_total + chick_total
    # Delivery stats combining chick and feed deliveries
    delivered_chicks = ChickRequest.objects.filter(delivered=True).count()
    delivered_feeds = FeedAllocation.objects.filter(delivered=True).count()
    deliveries_made = delivered_chicks + delivered_feeds
    pending_chick_deliv = ChickRequest.objects.filter(status='approved', delivered=False).count()
    pending_feed_deliv = FeedAllocation.objects.filter(status='approved', delivered=False).count()
    pending_deliveries = pending_chick_deliv + pending_feed_deliv
    farmers_with_feeds = FeedAllocation.objects.values('chick_request__farmer').distinct().count()
    farmers_paid = FeedAllocation.objects.filter(payment_status='paid').count()
    pending_payments = FeedAllocation.objects.filter(payment_status='pending').count()
    context = {
        'total_users': total_users,
        'chick_stock': chick_stock,
        'feed_stock': feed_stock,
    'completed_requests': deliveries_made,
        'approved_requests': approved_requests,
        'pending_requests': pending_requests,
        'rejected_requests': rejected_requests,
        'total_sales': total_sales,
        'deliveries_made': deliveries_made,
        'pending_deliveries': pending_deliveries,
        'farmers_with_feeds': farmers_with_feeds,
        'farmers_paid': farmers_paid,
        'pending_payments': pending_payments,
    }
    return render(request, 'managersdashboard.html', context)

@role_required('manager')
def ViewChickRequests(request):
    requests_qs = ChickRequest.objects.select_related('farmer').order_by('-request_date')[:200]
    return render(request, 'viewChickrequests.html', {'requests': requests_qs})

@role_required('manager')
def ViewFeedRequests(request):
    feed_allocs = FeedAllocation.objects.select_related('chick_request', 'chick_request__farmer', 'feed_stock').order_by('-id')[:200]
    return render(request, 'viewFeedAllocations.html', {'allocations': feed_allocs})

@role_required('manager')
def FarmerReview(request):
    requests = ChickRequest.objects.select_related('farmer').order_by('-request_date')
    return render(request, 'farmerReview.html', {'requests': requests})

@role_required('manager')
def Approverequest(request):
    return render(request, 'approveRequest.html')


@role_required('manager')
def FarmerRecords(request):
    farmers = Customer.objects.order_by('-registration_date')[:500]
    return render(request, 'farmerRecords.html', {'farmers': farmers})

@role_required('manager')
def chickStock(request):
    chick_stocks = ChickStock.objects.order_by('batch_name')
    return render(request, 'chickStock.html', {'chick_stocks': chick_stocks})

@role_required('manager')
def feedStock(request):
    feed_stocks = FeedStock.objects.order_by('stock_name')
    return render(request, 'feedStock.html', {'feed_stocks': feed_stocks})

@role_required('manager')
def UpdateChickStock(request, stock_id=None):
    # Try to get existing stock if ID is provided
    chickstock = None
    if stock_id:
        try:
            chickstock = ChickStock.objects.get(id=stock_id)
        except ChickStock.DoesNotExist:
            messages.error(request, "Stock not found!")
            return redirect('/chickstock/')
    
    # Handle form submission
    if request.method == 'POST':
        # Get form data
        batch_name = request.POST.get('batch_name')
        chick_type = request.POST.get('chick_type')
        chick_breed = request.POST.get('chick_breed')
    # chick_age removed from form per requirements; keep existing value when editing
        chick_price = request.POST.get('chick_price')
        stock_quantity = request.POST.get('stock_quantity')
        
        try:
            # Check if this batch already exists
            if chickstock:  # Editing existing stock
                existing_stock = chickstock
            else:
                existing_stock = ChickStock.objects.filter(batch_name=batch_name).first()
            
            if existing_stock:
                # Update existing stock
                existing_stock.batch_name = batch_name
                existing_stock.chick_type = chick_type
                existing_stock.chick_breed = chick_breed
                # preserve existing chick_age
                existing_stock.chick_price = chick_price
                existing_stock.stock_quantity = stock_quantity
                existing_stock.save()
                messages.success(request, f"Stock '{batch_name}' updated successfully!")
            else:
                # Create new stock entry
                new_stock = ChickStock(
                    batch_name=batch_name,
                    chick_type=chick_type,
                    chick_breed=chick_breed,
                    chick_age=1,  # default minimal age since field exists in model
                    chick_price=chick_price,
                    stock_quantity=stock_quantity
                )
                new_stock.save()
                messages.success(request, f"New stock '{batch_name}' added successfully!")
            
            # Redirect to the stock listing page
            return redirect('/chickstock/')
        
        except Exception as e:
            # Handle errors
            messages.error(request, f"Error updating stock: {str(e)}")
            
            # Re-populate the form with submitted values
            form_data = type('', (), {})()
            form_data.batch_name = batch_name
            form_data.chick_type = chick_type
            form_data.chick_breed = chick_breed
            # age not editable in UI
            form_data.chick_price = chick_price
            form_data.stock_quantity = stock_quantity
            return render(request, 'updateChickStock.html', {'chickstock': form_data})
    
    # For GET requests
    if not chickstock:
        # Show empty form for new stock
        class EmptyChickStock:
            def __init__(self):
                self.batch_name = ""
                self.chick_type = ""
                self.chick_breed = ""
                self.chick_age = 1
                self.chick_price = ""
                self.stock_quantity = ""
        
        chickstock = EmptyChickStock()
    
    return render(request, 'updateChickStock.html', {'chickstock': chickstock})

@role_required('manager')
def UpdateFeedStock(request, stock_id=None):
    # Try to get existing stock if ID is provided
    feed_stock = None
    if stock_id:
        try:
            feed_stock = FeedStock.objects.get(id=stock_id)
        except FeedStock.DoesNotExist:
            messages.error(request, 'Feed stock not found.')
            return redirect('Feedstock')
    
    if request.method == 'POST':
        try:
            if feed_stock:
                # Update existing feed stock
                feed_stock.stock_name = request.POST.get('stock_name')
                feed_stock.feed_name = request.POST.get('feed_name')
                feed_stock.feed_type = request.POST.get('feed_type')
                feed_stock.feed_brand = request.POST.get('feed_brand')
                feed_stock.feed_quantity = request.POST.get('feed_quantity')
                feed_stock.expiry_date = request.POST.get('expiry_date')
                feed_stock.purchase_price = request.POST.get('purchase_price')
                feed_stock.selling_price = request.POST.get('selling_price')
                feed_stock.supplier = request.POST.get('supplier')
                feed_stock.supplier_contact = request.POST.get('supplier_contact')
            else:
                # Create new feed stock
                feed_stock = FeedStock(
                    stock_name=request.POST.get('stock_name'),
                    feed_name=request.POST.get('feed_name'),
                    feed_type=request.POST.get('feed_type'),
                    feed_brand=request.POST.get('feed_brand'),
                    feed_quantity=request.POST.get('feed_quantity'),
                    expiry_date=request.POST.get('expiry_date'),
                    purchase_price=request.POST.get('purchase_price'),
                    selling_price=request.POST.get('selling_price'),
                    supplier=request.POST.get('supplier'),
                    supplier_contact=request.POST.get('supplier_contact')
                )
            
            feed_stock.full_clean()
            feed_stock.save()
            
            if stock_id:
                messages.success(request, 'Feed stock updated successfully!')
            else:
                messages.success(request, 'Feed stock added successfully!')
            return redirect('Feedstock')
        except Exception as e:
            messages.error(request, str(e))
    
    return render(request, 'updateFeedStock.html', {'feed_stock': feed_stock})

@role_required('manager')
def Reports(request):
    # Filters
    def parse_date(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except Exception:
            return None
    filters = {
        'start': request.GET.get('start') or '',
        'end': request.GET.get('end') or '',
        'chick_type': request.GET.get('chick_type') or '',
        'chick_breed': request.GET.get('chick_breed') or '',
        'feed_type': request.GET.get('feed_type') or '',
        'status': request.GET.get('status') or '',
        'agent': request.GET.get('agent') or '',
        'farmer': request.GET.get('farmer') or '',
        'q': request.GET.get('q') or '',
    }
    start_date = parse_date(filters['start'])
    end_date = parse_date(filters['end'])

    # Base querysets
    chick_requests_qs = ChickRequest.objects.select_related('farmer', 'created_by').order_by('-request_date')
    feed_allocations_qs = FeedAllocation.objects.select_related('chick_request', 'chick_request__farmer', 'feed_stock').order_by('-id')
    farmers_qs = Customer.objects.order_by('-registration_date')
    chick_stock_qs = ChickStock.objects.order_by('batch_name')
    feed_stock_qs = FeedStock.objects.order_by('stock_name')

    # Apply filters
    if start_date:
        chick_requests_qs = chick_requests_qs.filter(request_date__date__gte=start_date)
        feed_allocations_qs = feed_allocations_qs.filter(chick_request__request_date__date__gte=start_date)
        farmers_qs = farmers_qs.filter(registration_date__date__gte=start_date)
    if end_date:
        chick_requests_qs = chick_requests_qs.filter(request_date__date__lte=end_date)
        feed_allocations_qs = feed_allocations_qs.filter(chick_request__request_date__date__lte=end_date)
        farmers_qs = farmers_qs.filter(registration_date__date__lte=end_date)
    if filters['chick_type']:
        chick_requests_qs = chick_requests_qs.filter(chick_type=filters['chick_type'])
        chick_stock_qs = chick_stock_qs.filter(chick_type=filters['chick_type'])
    if filters['chick_breed']:
        chick_requests_qs = chick_requests_qs.filter(chick_breed=filters['chick_breed'])
        chick_stock_qs = chick_stock_qs.filter(chick_breed=filters['chick_breed'])
    if filters['feed_type']:
        feed_allocations_qs = feed_allocations_qs.filter(feed_type=filters['feed_type'])
        feed_stock_qs = feed_stock_qs.filter(feed_type=filters['feed_type'])
    if filters['status']:
        chick_requests_qs = chick_requests_qs.filter(status=filters['status'])
        feed_allocations_qs = feed_allocations_qs.filter(status=filters['status'])
    if filters['agent']:
        chick_requests_qs = chick_requests_qs.filter(created_by_id=filters['agent'])
        feed_allocations_qs = feed_allocations_qs.filter(chick_request__created_by_id=filters['agent'])
    if filters['farmer']:
        chick_requests_qs = chick_requests_qs.filter(farmer_id=filters['farmer'])
        feed_allocations_qs = feed_allocations_qs.filter(chick_request__farmer_id=filters['farmer'])
    if filters['q']:
        q = filters['q']
        chick_requests_qs = chick_requests_qs.filter(
            Q(chick_request_id__icontains=q) | Q(farmer__farmer_name__icontains=q) | Q(created_by__username__icontains=q)
        )
        feed_allocations_qs = feed_allocations_qs.filter(
            Q(feed_request_id__icontains=q) | Q(chick_request__chick_request_id__icontains=q) | Q(chick_request__farmer__farmer_name__icontains=q)
        )
        farmers_qs = farmers_qs.filter(Q(farmer_id__icontains=q) | Q(farmer_name__icontains=q))

    # Totals and stats
    chick_stock_total = chick_stock_qs.aggregate(total=Sum('stock_quantity'))['total'] or 0
    feed_stock_total = feed_stock_qs.aggregate(total=Sum('feed_quantity'))['total'] or 0

    # Sales totals
    feed_total = feed_allocations_qs.filter(status='approved').aggregate(
        total=Sum(F('bags_allocated') * F('feed_stock__selling_price'))
    )['total'] or 0
    chick_total = 0
    for req in chick_requests_qs.filter(status='approved'):
        price = ChickStock.objects.filter(
            chick_type=req.chick_type,
            chick_breed=req.chick_breed
        ).order_by('-updated_at').values_list('chick_price', flat=True).first() or 1650
        chick_total += int(req.quantity or 0) * int(price)
    total_sales = feed_total + chick_total

    # Use the same sliced datasets for tables and stats to keep summaries in sync with what's shown
    chick_requests_display = list(chick_requests_qs[:1000])
    feed_allocations_display = list(feed_allocations_qs[:1000])

    total_chick_requests = len(chick_requests_display)
    total_feed_allocations = len(feed_allocations_display)
    total_farmers = farmers_qs.count()
    pending_chick_requests = chick_requests_qs.filter(status='pending').count()
    pending_feed_payments = feed_allocations_qs.filter(payment_status='pending').count()
    low_stock_items = ChickStock.objects.filter(stock_quantity__lt=100).count() + FeedStock.objects.filter(feed_quantity__lt=50).count()

    # Status stats for block (compute from the same displayed lists)
    from collections import Counter
    sc_counter = Counter(getattr(r, 'status', '') for r in chick_requests_display)
    sf_counter = Counter(getattr(a, 'status', '') for a in feed_allocations_display)
    stats = {
        'cr_pending': sc_counter.get('pending', 0),
        'cr_approved': sc_counter.get('approved', 0),
        'cr_rejected': sc_counter.get('rejected', 0),
        'fa_pending': sf_counter.get('pending', 0),
        'fa_approved': sf_counter.get('approved', 0),
        'fa_rejected': sf_counter.get('rejected', 0),
    }

    # Activity charts (daily buckets) and weekly summary
    from collections import OrderedDict
    buckets = OrderedDict()
    def date_key(dt):
        d = dt.date() if hasattr(dt, 'date') else dt
        return d.strftime('%Y-%m-%d')
    for r in chick_requests_qs:
        k = date_key(r.request_date)
        buckets.setdefault(k, {'chicks':0,'feeds':0})
        buckets[k]['chicks'] += 1
    for a in feed_allocations_qs:
        k = date_key(a.chick_request.request_date)
        buckets.setdefault(k, {'chicks':0,'feeds':0})
        buckets[k]['feeds'] += 1
    activity_labels = list(buckets.keys())
    activity_chicks = [buckets[k]['chicks'] for k in activity_labels]
    activity_feeds = [buckets[k]['feeds'] for k in activity_labels]
    # Status mix for chicks (used by charts/summary); derive from same displayed counts
    status_mix = {k: sc_counter.get(k, 0) for k in ['pending','approved','rejected','completed']}

    # Weekly summary: ISO year-week
    from collections import defaultdict
    wk = defaultdict(lambda: {'chicks':0,'feeds':0})
    for r in chick_requests_qs:
        y, w, _ = r.request_date.isocalendar()
        wk[(y,w)]['chicks'] += 1
    for a in feed_allocations_qs:
        y, w, _ = a.chick_request.request_date.isocalendar()
        wk[(y,w)]['feeds'] += 1
    weekly_summary = [
        {
            'label': f"{y}-W{w:02d}",
            'chicks': v['chicks'],
            'feeds': v['feeds'],
            'total': v['chicks'] + v['feeds'],
        }
        for (y,w), v in sorted(wk.items())
    ]

    # Agent performance
    agent_perf_map = {}
    # Chick requests contribution
    for r in chick_requests_display:
        key = getattr(r.created_by, 'username', 'N/A')
        d = agent_perf_map.setdefault(key, {'total':0,'approved':0,'rejected':0,'delivered':0})
        d['total'] += 1
        if r.status=='approved': d['approved'] += 1
        if r.status=='rejected': d['rejected'] += 1
        if r.delivered: d['delivered'] += 1
    # Feed allocations contribution (by request owner)
    for a in feed_allocations_display:
        key = getattr(getattr(a.chick_request, 'created_by', None), 'username', 'N/A')
        d = agent_perf_map.setdefault(key, {'total':0,'approved':0,'rejected':0,'delivered':0})
        d['total'] += 1
        if a.status=='approved': d['approved'] += 1
        if a.status=='rejected': d['rejected'] += 1
        if a.delivered: d['delivered'] += 1
    agent_perf = [{'agent':k, **v} for k,v in sorted(agent_perf_map.items())]

    # Choices and aux lists
    chick_type_choices = ChickStock._meta.get_field('chick_type').choices
    chick_breed_choices = ChickStock._meta.get_field('chick_breed').choices
    feed_types = list(FeedStock.objects.values_list('feed_type', flat=True).distinct())
    agents = UserProfile.objects.filter(role='sales_agent').order_by('username')
    farmers_all = Customer.objects.order_by('farmer_name')

    # Trends (last 30 days vs previous 30), only when no explicit date filter
    trends = None
    if not start_date and not end_date:
        from datetime import timedelta
        today = timezone.now().date()
        cur_start = today - timedelta(days=30)
        prev_start = today - timedelta(days=60)
        prev_end = today - timedelta(days=30)
        # counts
        cr_cur = ChickRequest.objects.filter(request_date__date__gte=cur_start).count()
        cr_prev = ChickRequest.objects.filter(request_date__date__gte=prev_start, request_date__date__lt=prev_end).count()
        fa_cur = FeedAllocation.objects.filter(chick_request__request_date__date__gte=cur_start).count()
        fa_prev = FeedAllocation.objects.filter(chick_request__request_date__date__gte=prev_start, chick_request__request_date__date__lt=prev_end).count()
        # sales
        def calc_sales(start_d, end_d=None):
            feed_qs = FeedAllocation.objects.filter(status='approved', chick_request__request_date__date__gte=start_d)
            if end_d:
                feed_qs = feed_qs.filter(chick_request__request_date__date__lt=end_d)
            feed_t = feed_qs.aggregate(total=Sum(F('bags_allocated') * F('feed_stock__selling_price')))['total'] or 0
            chick_qs2 = ChickRequest.objects.filter(status='approved', request_date__date__gte=start_d)
            if end_d:
                chick_qs2 = chick_qs2.filter(request_date__date__lt=end_d)
            ct = 0
            for r in chick_qs2:
                price = ChickStock.objects.filter(chick_type=r.chick_type, chick_breed=r.chick_breed).order_by('-updated_at').values_list('chick_price', flat=True).first() or 1650
                ct += int(r.quantity or 0) * int(price)
            return (feed_t or 0) + (ct or 0)
        sales_cur = calc_sales(cur_start, None)
        sales_prev = calc_sales(prev_start, prev_end)
        def pct(cur, prev):
            try:
                return round(((cur - prev) / prev) * 100.0, 1) if prev else (100.0 if cur and not prev else 0.0)
            except Exception:
                return 0.0
        trends = {
            'sales': pct(sales_cur, sales_prev),
            'chick_requests': pct(cr_cur, cr_prev),
            'feed_allocations': pct(fa_cur, fa_prev),
        }

    context = {
        'filters': filters,
        'chick_stock': list(chick_stock_qs),
        'feed_stock': list(feed_stock_qs),
        'chick_stock_total': chick_stock_total,
        'feed_stock_total': feed_stock_total,
    'chick_requests': chick_requests_display,
    'feed_allocations': feed_allocations_display,
        'farmers': list(farmers_qs[:1000]),
        'stats': stats,
        'chick_type_choices': chick_type_choices,
        'chick_breed_choices': chick_breed_choices,
        'feed_types': feed_types,
        'agents': agents,
        'farmers_all': farmers_all,
        'total_sales': total_sales,
        'total_chick_requests': total_chick_requests,
        'total_feed_allocations': total_feed_allocations,
        'total_farmers': total_farmers,
        'pending_chick_requests': pending_chick_requests,
        'pending_feed_payments': pending_feed_payments,
        'low_stock_items': low_stock_items,
        'charts': {
            'activity_labels': activity_labels,
            'activity_chicks': activity_chicks,
            'activity_feeds': activity_feeds,
            'status_mix': status_mix,
        },
    'trends': trends,
    'weekly_summary': weekly_summary,
    'agent_perf': agent_perf,
    }
    return render(request, 'reports.html', context)

@role_required('manager')
def reports_export(request):
    dataset = request.GET.get('dataset')
    fmt = request.GET.get('format', 'csv')

    # Reuse Reports filters logic to build filtered querysets (without rendering)
    # We'll factor minimal parts here
    def parse_date(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except Exception:
            return None
    start = parse_date(request.GET.get('start') or '')
    end = parse_date(request.GET.get('end') or '')
    status = request.GET.get('status') or ''
    chick_type = request.GET.get('chick_type') or ''
    chick_breed = request.GET.get('chick_breed') or ''
    feed_type = request.GET.get('feed_type') or ''
    agent = request.GET.get('agent') or ''
    farmer = request.GET.get('farmer') or ''
    location = request.GET.get('location') or ''
    q = request.GET.get('q') or ''

    # Build datasets
    if dataset == 'chick_stock':
        qs = ChickStock.objects.all()
        if chick_type: qs = qs.filter(chick_type=chick_type)
        if chick_breed: qs = qs.filter(chick_breed=chick_breed)
        rows = [['Batch','Type','Breed','Age','Price','Qty']]
        for s in qs.order_by('batch_name'):
            rows.append([s.batch_name,s.chick_type,s.chick_breed,s.chick_age,s.chick_price,s.stock_quantity])
        filename = 'chick_stock'
    elif dataset == 'feed_stock':
        qs = FeedStock.objects.all()
        if feed_type: qs = qs.filter(feed_type=feed_type)
        rows = [['Stock','Feed','Type','Brand','Qty','Unit Price']]
        for s in qs.order_by('stock_name'):
            rows.append([s.stock_name,s.feed_name,s.feed_type,s.feed_brand,s.feed_quantity,s.selling_price])
        filename = 'feed_stock'
    elif dataset == 'chick_requests':
        qs = ChickRequest.objects.select_related('farmer','created_by')
        if start: qs = qs.filter(request_date__date__gte=start)
        if end: qs = qs.filter(request_date__date__lte=end)
        if status: qs = qs.filter(status=status)
        if chick_type: qs = qs.filter(chick_type=chick_type)
        if chick_breed: qs = qs.filter(chick_breed=chick_breed)
        if agent: qs = qs.filter(created_by_id=agent)
        if farmer: qs = qs.filter(farmer_id=farmer)
        if location: qs = qs.filter(farmer__location__icontains=location)
        if q: qs = qs.filter(Q(chick_request_id__icontains=q) | Q(farmer__farmer_name__icontains=q))
        rows = [['Req ID','Farmer','Type','Breed','Qty','Status','Date']]
        for r in qs.order_by('-request_date'):
            rows.append([r.chick_request_id, r.farmer.farmer_name, r.chick_type, r.chick_breed, r.quantity, r.status, r.request_date.strftime('%Y-%m-%d %H:%M')])
        filename = 'chick_requests'
    elif dataset == 'feed_allocations':
        qs = FeedAllocation.objects.select_related('chick_request','chick_request__farmer','feed_stock')
        if start: qs = qs.filter(chick_request__request_date__date__gte=start)
        if end: qs = qs.filter(chick_request__request_date__date__lte=end)
        if status: qs = qs.filter(status=status)
        if feed_type: qs = qs.filter(feed_type=feed_type)
        if agent: qs = qs.filter(chick_request__created_by_id=agent)
        if farmer: qs = qs.filter(chick_request__farmer_id=farmer)
        if location: qs = qs.filter(chick_request__farmer__location__icontains=location)
        if q: qs = qs.filter(Q(feed_request_id__icontains=q) | Q(chick_request__chick_request_id__icontains=q))
        rows = [['Feed ID','Req ID','Feed','Brand','Bags','Status','Payment']]
        for a in qs.order_by('-id'):
            rows.append([a.feed_request_id, a.chick_request.chick_request_id, a.feed_name, a.feed_brand, a.bags_allocated, a.status, a.payment_status])
        filename = 'feed_allocations'
    elif dataset == 'farmers':
        qs = Customer.objects.all()
        if start: qs = qs.filter(registration_date__date__gte=start)
        if end: qs = qs.filter(registration_date__date__lte=end)
        if location: qs = qs.filter(location__icontains=location)
        if q: qs = qs.filter(Q(farmer_id__icontains=q) | Q(farmer_name__icontains=q))
        rows = [['Farmer ID','Name','Gender','Age','Phone','Location','Registered']]
        for f in qs.order_by('-registration_date'):
            rows.append([f.farmer_id, f.farmer_name, f.gender, f.age, f.phone_number, f.location, f.registration_date.strftime('%Y-%m-%d %H:%M')])
        filename = 'farmers'
    elif dataset == 'agent_performance':
        # Aggregate from ChickRequest
        perf = {}
        qs = ChickRequest.objects.select_related('created_by')
        if start: qs = qs.filter(request_date__date__gte=start)
        if end: qs = qs.filter(request_date__date__lte=end)
        for r in qs:
            key = getattr(r.created_by, 'username', 'N/A')
            d = perf.setdefault(key, {'total':0,'approved':0,'rejected':0,'delivered':0})
            d['total'] += 1
            if r.status=='approved': d['approved'] += 1
            if r.status=='rejected': d['rejected'] += 1
            if r.delivered: d['delivered'] += 1
        rows = [['Agent','Requests','Approved','Rejected','Delivered']]
        for k,v in sorted(perf.items()):
            rows.append([k, v['total'], v['approved'], v['rejected'], v['delivered']])
        filename = 'agent_performance'
    elif dataset == 'activity_daily':
        # Build daily activity counts
        from collections import OrderedDict
        qsR = ChickRequest.objects.all()
        qsA = FeedAllocation.objects.all()
        if start: qsR = qsR.filter(request_date__date__gte=start)
        if end: qsR = qsR.filter(request_date__date__lte=end)
        if start: qsA = qsA.filter(chick_request__request_date__date__gte=start)
        if end: qsA = qsA.filter(chick_request__request_date__date__lte=end)
        buckets = OrderedDict()
        def key(d):
            return d.strftime('%Y-%m-%d')
        for r in qsR:
            buckets[key(r.request_date.date())] = buckets.get(key(r.request_date.date()), {'chicks':0,'feeds':0})
            buckets[key(r.request_date.date())]['chicks'] += 1
        for a in qsA:
            d = a.chick_request.request_date.date()
            buckets[key(d)] = buckets.get(key(d), {'chicks':0,'feeds':0})
            buckets[key(d)]['feeds'] += 1
        rows = [['Date','Chick Requests','Feed Allocations']]
        for k,v in buckets.items():
            rows.append([k, v['chicks'], v['feeds']])
        filename = 'activity_daily'
    elif dataset == 'activity_weekly':
        from collections import defaultdict
        qsR = ChickRequest.objects.all()
        qsA = FeedAllocation.objects.all()
        if start: qsR = qsR.filter(request_date__date__gte=start)
        if end: qsR = qsR.filter(request_date__date__lte=end)
        if start: qsA = qsA.filter(chick_request__request_date__date__gte=start)
        if end: qsA = qsA.filter(chick_request__request_date__date__lte=end)
        wk = defaultdict(lambda: {'chicks':0,'feeds':0})
        for r in qsR:
            y, w, _ = r.request_date.isocalendar()
            wk[(y,w)]['chicks'] += 1
        for a in qsA:
            y, w, _ = a.chick_request.request_date.isocalendar()
            wk[(y,w)]['feeds'] += 1
        rows = [['Year-Week','Chick Requests','Feed Allocations','Total']]
        for (y,w), v in sorted(wk.items()):
            rows.append([f"{y}-W{w:02d}", v['chicks'], v['feeds'], v['chicks']+v['feeds']])
        filename = 'activity_weekly'
    elif dataset == 'general':
        # Build a combined representation using same filters as the page
        # Reuse minimal parse/query bits from above
        # Chick stock
        cs = ChickStock.objects.all()
        if chick_type: cs = cs.filter(chick_type=chick_type)
        if chick_breed: cs = cs.filter(chick_breed=chick_breed)
        chick_stock_rows = [['Batch','Type','Breed','Age','Price','Qty']]
        for s in cs.order_by('batch_name'):
            chick_stock_rows.append([s.batch_name,s.chick_type,s.chick_breed,s.chick_age,s.chick_price,s.stock_quantity])
        # Feed stock
        fs = FeedStock.objects.all()
        if feed_type: fs = fs.filter(feed_type=feed_type)
        feed_stock_rows = [['Stock','Feed','Type','Brand','Qty','Unit Price']]
        for s in fs.order_by('stock_name'):
            feed_stock_rows.append([s.stock_name,s.feed_name,s.feed_type,s.feed_brand,s.feed_quantity,s.selling_price])
        # Chick requests
        cr = ChickRequest.objects.select_related('farmer','created_by')
        if start: cr = cr.filter(request_date__date__gte=start)
        if end: cr = cr.filter(request_date__date__lte=end)
        if status: cr = cr.filter(status=status)
        if chick_type: cr = cr.filter(chick_type=chick_type)
        if chick_breed: cr = cr.filter(chick_breed=chick_breed)
        if agent: cr = cr.filter(created_by_id=agent)
        if farmer: cr = cr.filter(farmer_id=farmer)
        if q: cr = cr.filter(Q(chick_request_id__icontains=q) | Q(farmer__farmer_name__icontains=q))
        chick_req_rows = [['Req ID','Farmer','Type','Breed','Qty','Status','Date']]
        for r in cr.order_by('-request_date'):
            chick_req_rows.append([r.chick_request_id, r.farmer.farmer_name, r.chick_type, r.chick_breed, r.quantity, r.status, r.request_date.strftime('%Y-%m-%d %H:%M')])
        # Feed allocations
        fa = FeedAllocation.objects.select_related('chick_request','chick_request__farmer','feed_stock')
        if start: fa = fa.filter(chick_request__request_date__date__gte=start)
        if end: fa = fa.filter(chick_request__request_date__date__lte=end)
        if status: fa = fa.filter(status=status)
        if feed_type: fa = fa.filter(feed_type=feed_type)
        if agent: fa = fa.filter(chick_request__created_by_id=agent)
        if farmer: fa = fa.filter(chick_request__farmer_id=farmer)
        if q: fa = fa.filter(Q(feed_request_id__icontains=q) | Q(chick_request__chick_request_id__icontains=q))
        feed_alloc_rows = [['Feed ID','Req ID','Feed','Brand','Bags','Status','Payment']]
        for a in fa.order_by('-id'):
            feed_alloc_rows.append([a.feed_request_id, a.chick_request.chick_request_id, a.feed_name, a.feed_brand, a.bags_allocated, a.status, a.payment_status])

        # Agent performance (aggregate across both datasets)
        perf = {}
        for r in cr:
            key = getattr(r.created_by, 'username', 'N/A')
            d = perf.setdefault(key, {'total':0,'approved':0,'rejected':0,'delivered':0})
            d['total'] += 1
            if r.status=='approved': d['approved'] += 1
            if r.status=='rejected': d['rejected'] += 1
            if r.delivered: d['delivered'] += 1
        for a in fa:
            key = getattr(getattr(a.chick_request, 'created_by', None), 'username', 'N/A')
            d = perf.setdefault(key, {'total':0,'approved':0,'rejected':0,'delivered':0})
            d['total'] += 1
            if a.status=='approved': d['approved'] += 1
            if a.status=='rejected': d['rejected'] += 1
            if a.delivered: d['delivered'] += 1
        agent_rows = [['Agent','Requests','Approved','Rejected','Delivered']]
        for k,v in sorted(perf.items()):
            agent_rows.append([k, v['total'], v['approved'], v['rejected'], v['delivered']])

        # Output formats
        if fmt == 'xlsx':
            try:
                from openpyxl import Workbook
            except Exception:
                return HttpResponse('Install openpyxl to enable Excel export: pip install openpyxl', content_type='text/plain')
            wb = Workbook()
            ws1 = wb.active; ws1.title = 'Chick Stock'
            for r in chick_stock_rows: ws1.append(r)
            ws2 = wb.create_sheet('Feed Stock')
            for r in feed_stock_rows: ws2.append(r)
            ws3 = wb.create_sheet('Chick Requests')
            for r in chick_req_rows: ws3.append(r)
            ws4 = wb.create_sheet('Feed Allocations')
            for r in feed_alloc_rows: ws4.append(r)
            ws5 = wb.create_sheet('Agent Performance')
            for r in agent_rows: ws5.append(r)
            from io import BytesIO
            buff = BytesIO(); wb.save(buff); buff.seek(0)
            resp = HttpResponse(buff.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = 'attachment; filename="general.xlsx"'
            return resp
        elif fmt == 'pdf':
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas
            except Exception:
                return HttpResponse('Install reportlab to enable PDF export: pip install reportlab', content_type='text/plain')
            from io import BytesIO
            buff = BytesIO(); c = canvas.Canvas(buff, pagesize=A4)
            width, height = A4; x0, y = 40, height-50
            def write_section(title, rows):
                nonlocal y
                c.setFont('Helvetica-Bold', 12); c.drawString(x0, y, title); y -= 18
                c.setFont('Helvetica', 10)
                for row in rows:
                    line = '  '.join(str(x) for x in row)
                    c.drawString(x0, y, line[:120]); y -= 14
                    if y < 60:
                        c.showPage(); y = height-50
            write_section('Chick Stock', chick_stock_rows)
            write_section('Feed Stock', feed_stock_rows)
            write_section('Chick Requests', chick_req_rows)
            write_section('Feed Allocations', feed_alloc_rows)
            write_section('Agent Performance', agent_rows)
            c.save(); buff.seek(0)
            resp = HttpResponse(buff.read(), content_type='application/pdf')
            resp['Content-Disposition'] = 'attachment; filename="general.pdf"'
            return resp
        else:
            return HttpResponseBadRequest('Unknown format')
    else:
        return HttpResponseBadRequest('Unknown dataset')

    # Writers per format
    if fmt == 'csv':
        import csv
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        writer = csv.writer(resp)
        for row in rows:
            writer.writerow(row)
        return resp
    elif fmt == 'xlsx':
        try:
            from openpyxl import Workbook
        except Exception:
            return HttpResponse('Install openpyxl to enable Excel export: pip install openpyxl', content_type='text/plain')
        wb = Workbook()
        ws = wb.active
        ws.title = 'Report'
        for r in rows:
            ws.append(r)
        from io import BytesIO
        buff = BytesIO()
        wb.save(buff)
        buff.seek(0)
        resp = HttpResponse(buff.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return resp
    elif fmt == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
        except Exception:
            return HttpResponse('Install reportlab to enable PDF export: pip install reportlab', content_type='text/plain')
        from io import BytesIO
        buff = BytesIO()
        c = canvas.Canvas(buff, pagesize=A4)
        width, height = A4
        x0, y = 40, height - 50
        for i,row in enumerate(rows):
            line = '  '.join(str(x) for x in row)
            c.drawString(x0, y, line[:120])
            y -= 16
            if y < 60:
                c.showPage(); y = height - 50
        c.save()
        buff.seek(0)
        resp = HttpResponse(buff.read(), content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return resp
    else:
        return HttpResponseBadRequest('Unknown format')

@role_required('manager')
def Sales(request):
    # Filters: optional start/end date (YYYY-MM-DD)
    start = request.GET.get('start')
    end = request.GET.get('end')
    def parse_date(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except Exception:
            return None
    start_date = parse_date(start)
    end_date = parse_date(end)

    # Feed sales rows: approved allocations with amounts = bags * selling_price
    feed_qs = FeedAllocation.objects.filter(status='approved').select_related('chick_request__farmer', 'feed_stock').order_by('-id')
    if start_date:
        feed_qs = feed_qs.filter(chick_request__request_date__date__gte=start_date)
    if end_date:
        feed_qs = feed_qs.filter(chick_request__request_date__date__lte=end_date)
    feed_rows = []
    feed_total = 0
    for a in feed_qs:
        price = getattr(a.feed_stock, 'selling_price', 0) or 0
        amount = int(a.bags_allocated or 0) * int(price)
        feed_total += amount
        feed_rows.append({
            'feed_request_id': a.feed_request_id,
            'req_id': a.chick_request.chick_request_id,
            'farmer': a.chick_request.farmer.farmer_name,
            'feed_name': a.feed_name,
            'brand': a.feed_brand,
            'bags': a.bags_allocated,
            'unit_price': price,
            'amount': amount,
            'date': a.chick_request.request_date,
        })

    # Chick sales rows: approved chick requests with amount = qty * price (latest for type/breed)
    chick_qs = ChickRequest.objects.filter(status='approved').select_related('farmer').order_by('-request_date')
    if start_date:
        chick_qs = chick_qs.filter(request_date__date__gte=start_date)
    if end_date:
        chick_qs = chick_qs.filter(request_date__date__lte=end_date)
    chick_rows = []
    chick_total = 0
    for r in chick_qs:
        price = ChickStock.objects.filter(
            chick_type=r.chick_type,
            chick_breed=r.chick_breed
        ).order_by('-updated_at').values_list('chick_price', flat=True).first() or 1650
        amount = int(r.quantity or 0) * int(price)
        chick_total += amount
        chick_rows.append({
            'req_id': r.chick_request_id,
            'farmer': r.farmer.farmer_name,
            'type': r.chick_type,
            'breed': r.chick_breed,
            'qty': r.quantity,
            'unit_price': price,
            'amount': amount,
            'date': r.request_date,
        })

    total_sales = feed_total + chick_total

    return render(request, 'sales.html', {
        'feed_rows': feed_rows,
        'feed_total': feed_total,
        'chick_rows': chick_rows,
        'chick_total': chick_total,
        'total_sales': total_sales,
        'start': start or '',
        'end': end or '',
    })



@role_required('manager')
def DeleteRequest(request):
    return render(request, 'deleteRequest.html')

@role_required('manager')
def ApproveFeedRequest(request, request_id):
    try:
        feed_allocation = FeedAllocation.objects.select_related('feed_stock').get(id=request_id)

        if request.method == 'POST':
            action = request.POST.get('action')

            if action not in ('approve', 'reject'):
                messages.error(request, 'Invalid action.')
                return redirect('Viewfeedrequests')

            with transaction.atomic():
                if action == 'approve':
                    # Ensure linked stock and enough quantity
                    stock = feed_allocation.feed_stock
                    bags = int(feed_allocation.bags_allocated or 0)
                    if stock and (stock.feed_quantity or 0) >= bags:
                        # Deduct inventory
                        stock.feed_quantity = (stock.feed_quantity or 0) - bags
                        stock.save(update_fields=['feed_quantity'])
                        # Update allocation status only (do not force payment status)
                        feed_allocation.status = 'approved'
                        feed_allocation.save(update_fields=['status'])
                        messages.success(request, 'Feed request approved successfully!')
                    else:
                        messages.error(request, 'Insufficient stock to approve this request.')
                        return redirect('Viewfeedrequests')
                else:  # reject
                    feed_allocation.status = 'rejected'
                    feed_allocation.payment_status = 'rejected'
                    feed_allocation.save(update_fields=['status', 'payment_status'])
                    messages.success(request, 'Feed request rejected successfully!')

        return redirect('Viewfeedrequests')

    except FeedAllocation.DoesNotExist:
        messages.error(request, 'Feed request not found.')
        return redirect('Viewfeedrequests')




#For the Sales Agents
@role_required('sales_agent')
def SalesAgentdashboard(request):
    my_requests = ChickRequest.objects.filter(created_by=request.user)
    status_counts = my_requests.values('status').annotate(c=Count('id'))
    sc_map = {s['status']: s['c'] for s in status_counts}
    total_my_requests = my_requests.count()
    total_chicks_requested = my_requests.aggregate(total=Sum('quantity'))['total'] or 0
    # Count deliveries across chicks and feed allocations for this agent
    delivered_chicks = my_requests.filter(delivered=True).count()
    delivered_feeds = FeedAllocation.objects.filter(chick_request__created_by=request.user, delivered=True).count()
    delivered = delivered_chicks + delivered_feeds
    context = {
        'total_my_requests': total_my_requests,
        'total_chicks_requested': total_chicks_requested,
        'my_pending': sc_map.get('pending', 0),
        'my_approved': sc_map.get('approved', 0),
        'my_rejected': sc_map.get('rejected', 0),
        'my_completed': sc_map.get('completed', 0),
        'delivered': delivered,
    }
    return render(request, '1salesAgentdashboard.html', context)


@role_required('sales_agent')
def ViewSalesAgentFarmers(request):
    from django.db.models import Q
    # Include both new linkage via sales_agent and legacy via registered_by username
    base_qs = Customer.objects.filter(
        Q(sales_agent=request.user) | Q(registered_by=request.user.username)
    ).select_related('sales_agent').order_by('-registration_date')

    search_query = request.GET.get('q', '').strip()
    if search_query:
        base_qs = base_qs.filter(
            Q(farmer_name__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(farmer_id__icontains=search_query)
        )

    return render(request, '1viewfarmers.html', {'farmers': base_qs})

@role_required('sales_agent')
def EditFarmer(request, farmer_id):
    from django.db.models import Q
    farmer_instance = get_object_or_404(Customer, 
        Q(id=farmer_id) & (Q(sales_agent=request.user) | Q(registered_by=request.user.username)))

    if request.method == 'POST':
        try:
            # Update farmer_instance fields from POST data
            farmer_instance.farmer_name = request.POST.get('farmer_name')
            farmer_instance.date_of_birth = datetime.strptime(request.POST.get('date_of_birth'), '%Y-%m-%d').date()
            farmer_instance.gender = request.POST.get('gender')
            farmer_instance.location = request.POST.get('location')
            farmer_instance.nin = request.POST.get('nin')
            farmer_instance.phone_number = request.POST.get('phone_number')
            farmer_instance.recommender_name = request.POST.get('recommender_name')
            farmer_instance.recommender_nin = request.POST.get('recommender_nin')
            farmer_instance.recommender_tel = request.POST.get('recommender_tel')

            # Validate age based on date of birth
            today = timezone.now().date()
            age = today.year - farmer_instance.date_of_birth.year - ((today.month, today.day) < (farmer_instance.date_of_birth.month, farmer_instance.date_of_birth.day))
            if age < 20 or age > 30:
                raise ValueError('Farmer must be between 20 and 30 years old.')

            # Validate phone number and recommender tel length (exactly 10 characters)
            if len(farmer_instance.phone_number) != 10:
                raise ValueError('Phone Number must be exactly 10 characters.')
            if len(farmer_instance.recommender_tel) != 10:
                raise ValueError('Recommender Tel must be exactly 10 characters.')
            
            # Validate recommender NIN (exactly 14 characters)
            if len(farmer_instance.recommender_nin) != 14:
                raise ValueError('Recommender NIN must be exactly 14 characters.')

            farmer_instance.full_clean()
            farmer_instance.save()
            messages.success(request, 'Farmer details updated successfully!')
            return redirect('salesagentfarmers')
        except Exception as e:
            messages.error(request, str(e))
    # Calculate gender selection flags to avoid template syntax issues
    is_male = farmer_instance.gender == 'M'
    is_female = farmer_instance.gender == 'F'
    
    return render(request, 'editFarmer.html', {
        'farmer': farmer_instance,
        'is_male': is_male,
        'is_female': is_female
    })

@role_required('sales_agent')
def DeleteFarmer(request, farmer_id):
    from django.db.models import Q
    farmer = get_object_or_404(Customer, 
        Q(id=farmer_id) & (Q(sales_agent=request.user) | Q(registered_by=request.user.username)))
    if request.method == 'POST':
        try:
            farmer.delete()
            messages.success(request, 'Farmer deleted successfully.')
            return redirect('salesagentfarmers')
        except Exception as e:
            messages.error(request, str(e))
    return render(request, 'deleteFarmer.html', {'farmer': farmer})

@role_required('sales_agent')
def RegisterFarmer(request):
    if request.method == 'POST':
        try:
            # Generate a unique username for the farmer
            import uuid
            username = f"farmer_{uuid.uuid4().hex[:8]}"
            
            # Validate age based on date of birth
            date_of_birth_str = request.POST.get('date_of_birth')
            date_of_birth = datetime.strptime(date_of_birth_str, '%Y-%m-%d').date()
            today = timezone.now().date()
            age = today.year - date_of_birth.year - ((today.month, today.day) < (date_of_birth.month, date_of_birth.day))
            
            if age < 20 or age > 30:
                raise ValueError('Farmer must be between 20 and 30 years old.')
            
            # Create user profile with farmer role
            user = UserProfile.objects.create_user(
                username=username,
                password=username,  # initial password (could be improved)
                role='farmer'
            )
            
            # Validate phone number and recommender tel length (exactly 10 characters)
            phone_number = request.POST.get('phone_number')
            recommender_tel = request.POST.get('recommender_tel')
            recommender_nin = request.POST.get('recommender_nin')
            
            if len(phone_number) != 10:
                raise ValueError('Phone Number must be exactly 10 characters.')
            if len(recommender_tel) != 10:
                raise ValueError('Recommender Tel must be exactly 10 characters.')
            
            # Validate recommender NIN (exactly 14 characters)
            if len(recommender_nin) != 14:
                raise ValueError('Recommender NIN must be exactly 14 characters.')
            
            # Create customer record (farmer_id will be auto-generated)
            customer = Customer(
                user=user,
                farmer_name=request.POST.get('farmer_name'),
                date_of_birth=date_of_birth,
                gender=request.POST.get('gender'),
                location=request.POST.get('location'),
                nin=request.POST.get('nin'),
                phone_number=phone_number,
                recommender_name=request.POST.get('recommender_name'),
                recommender_nin=request.POST.get('recommender_nin'),
                recommender_tel=recommender_tel,
                registered_by=request.user.username,
                sales_agent=request.user # Assign the current sales agent
            )
            customer.full_clean()
            customer.save()
            messages.success(request, f'Farmer registered successfully! Farmer ID: {customer.farmer_id}')
            return redirect('salesagentdashboard')
        except Exception as e:
            messages.error(request, str(e))
    return render(request, '1registerfarmer.html')

@role_required('sales_agent')
def AddChickRequest(request):
    if request.method == 'POST':
        try:
            farmer_id = request.POST.get('farmer')
            farmer = get_object_or_404(Customer, id=farmer_id)
            farmer_type = request.POST.get('farmer_type')
            quantity = int(request.POST.get('quantity', 0))
            
            # Validation for farmer type and quantity
            if farmer_type == 'starter' and quantity > 100:
                raise ValueError('Starter farmers cannot request more than 100 chicks.')
            if farmer_type == 'returning' and quantity > 500:
                raise ValueError('Returning farmers cannot request more than 500 chicks.')
            
            chick_request = ChickRequest(
                farmer=farmer,
                farmer_type=farmer_type,
                chick_type=request.POST.get('chick_type'),
                chick_breed=request.POST.get('chick_breed'),
                quantity=quantity,
                chick_period=1,  # field kept in model; default to 1 day since input was removed
                feed_taken=request.POST.get('feed_taken') == 'True',
                feed_name=request.POST.get('feed_name') or None,
                payment_terms=request.POST.get('payment_terms'),
                received_through=request.POST.get('received_through'),
                created_by=request.user
            )
            chick_request.full_clean()
            chick_request.save()
            messages.success(request, f'Chick request submitted successfully! Request ID: {chick_request.chick_request_id}')
            return redirect('salesagentdashboard')
        except Exception as e:
            messages.error(request, str(e))
    
    # Get all farmers for the dropdown
    farmers = Customer.objects.all().order_by('farmer_name')
    # Approved feeds for this agent
    approved_feeds = FeedAllocation.objects.filter(
        chick_request__created_by=request.user,
        status='approved'
    ).order_by('-id')
    # Build stock counts for dynamic breed display
    def count_for(t, b):
        return sum(s.stock_quantity or 0 for s in ChickStock.objects.filter(chick_type=t, chick_breed=b))
    stock_counts = {
        'layer': {'local': count_for('layer', 'local'), 'exotic': count_for('layer', 'exotic')},
        'broiler': {'local': count_for('broiler', 'local'), 'exotic': count_for('broiler', 'exotic')},
    }
    return render(request, '1addChickRequests.html', {
        'farmers': farmers,
        'approved_feeds': approved_feeds,
        'stock_counts': stock_counts,
    })

@role_required('sales_agent')
def AddFeedRequest(request):
    # Get chick requests belonging to this agent (approved or pending)
    chick_requests = ChickRequest.objects.filter(created_by=request.user, status__in=['approved', 'pending']).select_related('farmer')
    
    # Get available feed stock
    feed_stocks = FeedStock.objects.filter(feed_quantity__gt=0).order_by('feed_name')
    
    if request.method == 'POST':
        try:
            feed_stock_id = request.POST.get('feed_stock') or None
            feed_stock = FeedStock.objects.filter(id=feed_stock_id).first() if feed_stock_id else None
            
            feed_allocation = FeedAllocation(
                feed_stock=feed_stock,
                feed_name=feed_stock.feed_name,
                feed_type=feed_stock.feed_type,
                feed_brand=feed_stock.feed_brand,
                chick_request=get_object_or_404(ChickRequest, id=request.POST.get('chick_request')),
                bags_allocated=request.POST.get('bags_allocated'),
                amount_due=request.POST.get('amount_due'),
                payment_due_date=request.POST.get('payment_due_date'),
                payment_status=request.POST.get('payment_status')
            )
            feed_allocation.full_clean()
            feed_allocation.save()
            messages.success(request, f'Feed request submitted successfully! Feed Request ID: {feed_allocation.feed_request_id}')
            return redirect('salesagentdashboard')
        except Exception as e:
            messages.error(request, str(e))
    
    return render(request, '1addFeedRequest.html', {
        'chick_requests': chick_requests,
        'feed_stocks': feed_stocks
    })

@role_required('sales_agent')
def ViewSalesAgentChickRequests(request):
    my_requests = ChickRequest.objects.filter(created_by=request.user).select_related('farmer').order_by('-request_date')
    return render(request, '1viewchickrequests.html', {'chick_requests': my_requests})

@role_required('sales_agent')
def ViewSalesAgentFeedRequests(request):
    my_feed = FeedAllocation.objects.filter(chick_request__created_by=request.user).select_related('chick_request', 'chick_request__farmer', 'feed_stock').order_by('-id')
    return render(request, '1viewfeedrequests.html', {'allocations': my_feed})


@role_required('manager')
def ApproveChickRequest(request, request_id: int):
    """Approve or reject a chick request; on approval, deduct chick stock by type/breed."""
    req = get_object_or_404(ChickRequest, id=request_id)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action not in ('approve', 'reject'):
            messages.error(request, 'Invalid action.')
            return redirect('Viewchickrequests')
        if action == 'reject':
            req.status = 'rejected'
            req.save(update_fields=['status'])
            messages.success(request, 'Chick request rejected successfully!')
            return redirect('Viewchickrequests')
        # approve
        with transaction.atomic():
            # Sum available stock for the requested type/breed
            stocks = list(ChickStock.objects.select_for_update().filter(
                chick_type=req.chick_type,
                chick_breed=req.chick_breed
            ).order_by('-stock_quantity'))
            needed = int(req.quantity or 0)
            available = sum(s.stock_quantity or 0 for s in stocks)
            if available < needed:
                messages.error(request, 'Insufficient chick stock for the requested type/breed.')
                return redirect('Viewchickrequests')
            remaining = needed
            for s in stocks:
                if remaining <= 0:
                    break
                take = min(remaining, s.stock_quantity or 0)
                s.stock_quantity = (s.stock_quantity or 0) - take
                s.save(update_fields=['stock_quantity'])
                remaining -= take
            req.status = 'approved'
            req.approved_on = timezone.now()
            req.save(update_fields=['status', 'approved_on'])
            messages.success(request, 'Chick request approved and stock updated!')
    return redirect('Viewchickrequests')


# --- TXT Export endpoints (manager) ---
@role_required('manager')
def export_sales_txt(request):
    # Build combined sales export from approved feed allocations and chick requests
    lines = ['TYPE\tREQ_ID\tFARMER\tITEM\tQTY/BAGS\tUNIT_PRICE\tAMOUNT\tDATE\n']
    # Feed
    for a in FeedAllocation.objects.filter(status='approved').select_related('chick_request__farmer', 'feed_stock').order_by('id'):
        price = getattr(a.feed_stock, 'selling_price', 0) or 0
        amount = int(a.bags_allocated or 0) * int(price)
        lines.append(f"FEED\t{a.chick_request.chick_request_id}\t{a.chick_request.farmer.farmer_name}\t{a.feed_name}\t{a.bags_allocated}\t{price}\t{amount}\t{a.chick_request.request_date.date()}\n")
    # Chicks
    for r in ChickRequest.objects.filter(status='approved').select_related('farmer').order_by('request_date'):
        price = ChickStock.objects.filter(chick_type=r.chick_type, chick_breed=r.chick_breed).order_by('-updated_at').values_list('chick_price', flat=True).first() or 1650
        amount = int(r.quantity or 0) * int(price)
        lines.append(f"CHICKS\t{r.chick_request_id}\t{r.farmer.farmer_name}\t{r.chick_type}/{r.chick_breed}\t{r.quantity}\t{price}\t{amount}\t{r.request_date.date()}\n")
    resp = HttpResponse(''.join(lines), content_type='text/plain')
    resp['Content-Disposition'] = 'attachment; filename="sales.txt"'
    return resp

@role_required('manager')
def export_chick_requests_txt(request):
    header = 'REQ_ID\tFARMER\tTYPE\tBREED\tQTY\tSTATUS\tDATE\n'
    rows = [
        f"{r.chick_request_id}\t{r.farmer.farmer_name}\t{r.chick_type}\t{r.chick_breed}\t{r.quantity}\t{r.status}\t{r.request_date.date()}\n"
        for r in ChickRequest.objects.select_related('farmer').order_by('request_date')
    ]
    resp = HttpResponse(header + ''.join(rows), content_type='text/plain')
    resp['Content-Disposition'] = 'attachment; filename="chick_requests.txt"'
    return resp

@role_required('manager')
def export_feed_allocations_txt(request):
    header = 'FEED_ID\tREQ_ID\tFEED\tBAGS\tSTATUS\tPAYMENT\n'
    rows = [
        f"{a.feed_request_id}\t{a.chick_request.chick_request_id}\t{a.feed_name}\t{a.bags_allocated}\t{a.status}\t{a.payment_status}\n"
        for a in FeedAllocation.objects.select_related('chick_request').order_by('id')
    ]
    resp = HttpResponse(header + ''.join(rows), content_type='text/plain')
    resp['Content-Disposition'] = 'attachment; filename="feed_allocations.txt"'
    return resp

@role_required('manager')
def export_chick_stock_txt(request):
    header = 'BATCH\tTYPE\tBREED\tAGE\tPRICE\tQTY\n'
    rows = [
        f"{s.batch_name}\t{s.chick_type}\t{s.chick_breed}\t{s.chick_age}\t{s.chick_price}\t{s.stock_quantity}\n"
        for s in ChickStock.objects.order_by('batch_name')
    ]
    resp = HttpResponse(header + ''.join(rows), content_type='text/plain')
    resp['Content-Disposition'] = 'attachment; filename="chick_stock.txt"'
    return resp

@role_required('manager')
def export_feed_stock_txt(request):
    header = 'STOCK\tFEED\tTYPE\tBRAND\tQTY\tPRICE\n'
    rows = [
        f"{s.stock_name}\t{s.feed_name}\t{s.feed_type}\t{s.feed_brand}\t{s.feed_quantity}\t{s.selling_price}\n"
        for s in FeedStock.objects.order_by('stock_name')
    ]
    resp = HttpResponse(header + ''.join(rows), content_type='text/plain')
    resp['Content-Disposition'] = 'attachment; filename="feed_stock.txt"'
    return resp

@role_required('manager')
def export_farmers_txt(request):
    header = 'FARMER_ID\tNAME\tGENDER\tAGE\tPHONE\tLOCATION\n'
    rows = [
        f"{c.farmer_id}\t{c.farmer_name}\t{c.gender}\t{c.age}\t{c.phone_number}\t{c.location}\n"
        for c in Customer.objects.order_by('farmer_id')
    ]
    resp = HttpResponse(header + ''.join(rows), content_type='text/plain')
    resp['Content-Disposition'] = 'attachment; filename="farmers.txt"'
    return resp


def Logout(request):
    logout(request)
    return redirect('/')

    return render(request, '1viewfeedrequests.html', {'allocations': my_feed})





def Logout(request):

    logout(request)

    return redirect('/')


# --- Deliveries Management (manager) ---
@role_required('manager')
def Deliveries(request):
    approved_chicks = ChickRequest.objects.filter(status='approved').select_related('farmer').order_by('-approved_on')
    approved_feeds = FeedAllocation.objects.filter(status='approved').select_related('chick_request__farmer').order_by('-id')
    return render(request, 'deliveries.html', {
        'approved_chicks': approved_chicks,
        'approved_feeds': approved_feeds,
    })

@role_required('manager')
def MarkChickDelivered(request, req_id:int):
    req = get_object_or_404(ChickRequest, id=req_id)
    if request.method == 'POST':
        req.delivered = True
        req.status = req.status  # keep status as-approved; deliveries counted separately
        req.save(update_fields=['delivered'])
        messages.success(request, f"Marked {req.chick_request_id} as delivered.")
    return redirect('Deliveries')

@role_required('manager')
def MarkFeedDelivered(request, alloc_id:int):
    alloc = get_object_or_404(FeedAllocation, id=alloc_id)
    if request.method == 'POST':
        alloc.delivered = True
        alloc.save(update_fields=['delivered'])
        messages.success(request, f"Marked {alloc.feed_request_id} as delivered.")
    return redirect('Deliveries')


